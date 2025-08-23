import os, pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter
from config import REPORT_CONFIG, EXCEL_STYLES

def aplicar_formato(ws, filas, columnas):
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for row in ws.iter_rows(min_row=filas[0], max_row=filas[1], min_col=columnas[0], max_col=columnas[1]):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for col in range(columnas[0], columnas[1] + 1):
        cell = ws.cell(row=filas[0], column=col)
        cell.font = Font(bold=True, color=EXCEL_STYLES["HEADER_FONT_COLOR"])
        cell.fill = PatternFill(start_color=EXCEL_STYLES["HEADER_COLOR"], end_color=EXCEL_STYLES["HEADER_COLOR"], fill_type="solid")
        # Ajustar ancho según la columna
        if col == columnas[0]:  # Primera columna (ID/Servicio)
            ws.column_dimensions[get_column_letter(col)].width = 15
        elif col == columnas[0] + 1:  # Segunda columna (Nombre/descripción)
            ws.column_dimensions[get_column_letter(col)].width = 25
        elif col == columnas[0] + 2:  # Tercera columna (Métrica)
            ws.column_dimensions[get_column_letter(col)].width = 20
        elif col == columnas[0] + 3:  # Cuarta columna (Servicio)
            ws.column_dimensions[get_column_letter(col)].width = 30
        else:  # Columnas de números
            ws.column_dimensions[get_column_letter(col)].width = 12

def generar_excel(df, resumen, periodo, horas=None, resumen_servicio=pd.DataFrame()):
    try:
        sufijo = f"_ultimas_{horas}h" if horas else ""
        archivo = f'{REPORT_CONFIG["EXCEL_DIR"]}/Alertas_{periodo}{sufijo}.xlsx'
        os.makedirs(REPORT_CONFIG["EXCEL_DIR"], exist_ok=True)

        with pd.ExcelWriter(archivo, engine='openpyxl') as writer:
            # Detalle
            if not df.empty:
                columnas = ['Id cuenta', 'Nombre cuenta', 'Metrica', 'Servicio', 'Estado', 'Fecha']
                for col in ['Región', 'Motivo', 'Namespace']:
                    if col in df.columns and df[col].notna().any():
                        columnas.append(col)
                df_detalle = df[columnas].rename(columns={'Id cuenta': 'ID Cuenta', 'Nombre cuenta': 'Nombre Cuenta', 'Metrica': 'Métrica', 'Servicio': 'Servicio/Recurso'})
                df_detalle.to_excel(writer, sheet_name='Detalle', index=False)
                aplicar_formato(writer.sheets['Detalle'], [1, len(df_detalle) + 1], [1, len(df_detalle.columns)])
            else:
                pd.DataFrame({'Mensaje': ['No hay alertas en el período seleccionado']}).to_excel(writer, sheet_name='Detalle', index=False)
            
            # Resumen
            pd.DataFrame().to_excel(writer, sheet_name='Resumen', index=False)
            ws2 = writer.sheets['Resumen']
            
            # KPIs
            total, criticas, warnings, info = len(df), len(df[df['Estado'] == 'Critica']), len(df[df['Estado'] == 'Warning']), len(df[df['Estado'] == 'Informativo'])
            pct_criticas, pct_warnings, pct_info = (criticas / total * 100) if total > 0 else 0, (warnings / total * 100) if total > 0 else 0, (info / total * 100) if total > 0 else 0
            
            ws2['B2'] = "DASHBOARD DE ALERTAS"
            ws2['B2'].font = Font(bold=True, size=16)
            ws2.merge_cells('B2:F2')
            ws2['B2'].alignment = Alignment(horizontal='center')
            
            ws2['B4'], ws2['C4'], ws2['D4'], ws2['E4'] = "Total Alertas", "Críticas", "Warnings", "Informativas"
            ws2['B5'], ws2['C5'], ws2['D5'], ws2['E5'] = total, criticas, warnings, info
            ws2['C6'], ws2['D6'], ws2['E6'] = f"{pct_criticas:.1f}%", f"{pct_warnings:.1f}%", f"{pct_info:.1f}%"
            
            ws2['B5'].font = Font(bold=True, size=14)
            ws2['C5'].font = Font(bold=True, size=14, color=EXCEL_STYLES["CRITICAL_COLOR"])
            ws2['D5'].font = Font(bold=True, size=14, color=EXCEL_STYLES["WARNING_COLOR"])
            ws2['E5'].font = Font(bold=True, size=14, color=EXCEL_STYLES["INFO_COLOR"])
            
            for col in ['B', 'C', 'D', 'E']:
                ws2[f'{col}4'].font = Font(bold=True)
                for row in range(4, 7):
                    ws2.cell(row=row, column=ord(col)-64).alignment = Alignment(horizontal='center')
                    ws2.cell(row=row, column=ord(col)-64).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            
            # Resumen por cuenta
            if not resumen.empty:
                ws2.cell(row=13, column=2, value="RESUMEN POR CUENTA")
                ws2.cell(row=13, column=2).font = Font(bold=True, size=14)
                ws2.merge_cells(start_row=13, start_column=2, end_row=13, end_column=9)
                ws2.cell(row=13, column=2).alignment = Alignment(horizontal='center')
                
                # Headers dinámicos basados en las columnas del resumen
                headers = ['ID Cuenta', 'Nombre Cuenta', 'Métrica', 'Servicio/Recurso']
                for col in resumen.columns[4:]:  # Agregar solo las columnas de estado que existen
                    if col == 'Total':
                        headers.append('Total')
                    else:
                        headers.append(col)
                
                for j, header in enumerate(headers):
                    ws2.cell(row=15, column=2+j, value=header)
                
                for i, row in enumerate(resumen.itertuples(index=False)):
                    for j, val in enumerate(row):
                        ws2.cell(row=16+i, column=2+j, value=val)
                
                aplicar_formato(ws2, [15, 15+len(resumen)], [2, 9])
            
            # Resumen por servicio
            if not resumen_servicio.empty:
                fila_inicio = 17 + len(resumen) if not resumen.empty else 13
                ws2.cell(row=fila_inicio, column=2, value="RESUMEN POR SERVICIO")
                ws2.cell(row=fila_inicio, column=2).font = Font(bold=True, size=14)
                ws2.merge_cells(start_row=fila_inicio, start_column=2, end_row=fila_inicio, end_column=6)
                ws2.cell(row=fila_inicio, column=2).alignment = Alignment(horizontal='center')
                
                headers_servicio = ['Servicio/Recurso'] + [col for col in resumen_servicio.columns if col != 'Servicio']
                for j, header in enumerate(headers_servicio):
                    ws2.cell(row=fila_inicio+2, column=2+j, value=header)
                
                for i, row in enumerate(resumen_servicio.itertuples(index=False)):
                    for j, val in enumerate(row):
                        ws2.cell(row=fila_inicio+3+i, column=2+j, value=val)
                
                aplicar_formato(ws2, [fila_inicio+2, fila_inicio+2+len(resumen_servicio)], [2, 2+len(headers_servicio)-1])
                
                # Gráfico
                if criticas + warnings + info > 0:
                    try:
                        chart_data = [("Críticas", criticas), ("Warnings", warnings), ("Informativas", info)]
                        chart_data = [(label, value) for label, value in chart_data if value > 0]
                        
                        fila_grafico = fila_inicio + 5 + len(resumen_servicio)
                        ws2.cell(row=fila_grafico, column=2, value="Estado")
                        ws2.cell(row=fila_grafico, column=3, value="Cantidad")
                        
                        for i, (label, value) in enumerate(chart_data):
                            ws2.cell(row=fila_grafico+1+i, column=2, value=label)
                            ws2.cell(row=fila_grafico+1+i, column=3, value=value)
                        
                        aplicar_formato(ws2, [fila_grafico, fila_grafico+len(chart_data)], [2, 3])
                        
                        chart = PieChart()
                        chart.title = "Distribución de Alertas"
                        chart.height, chart.width = EXCEL_STYLES["CHART_HEIGHT"], EXCEL_STYLES["CHART_WIDTH"]
                        
                        data_ref = Reference(ws2, min_col=3, max_col=3, min_row=fila_grafico+1, max_row=fila_grafico+len(chart_data))
                        labels_ref = Reference(ws2, min_col=2, min_row=fila_grafico+1, max_row=fila_grafico+len(chart_data))
                        
                        chart.add_data(data_ref)
                        chart.set_categories(labels_ref)
                        
                        chart.dataLabels = DataLabelList()
                        chart.dataLabels.showPercent = True
                        
                        ws2.add_chart(chart, f"E{fila_grafico+len(chart_data)+3}")
                    except: pass
            else:
                # Gráfico cuando no hay resumen por servicio
                if criticas + warnings + info > 0:
                    try:
                        chart_data = [("Críticas", criticas), ("Warnings", warnings), ("Informativas", info)]
                        chart_data = [(label, value) for label, value in chart_data if value > 0]
                        
                        ws2.cell(row=18+len(resumen), column=2, value="Estado")
                        ws2.cell(row=18+len(resumen), column=3, value="Cantidad")
                        
                        for i, (label, value) in enumerate(chart_data):
                            ws2.cell(row=19+len(resumen)+i, column=2, value=label)
                            ws2.cell(row=19+len(resumen)+i, column=3, value=value)
                        
                        aplicar_formato(ws2, [18+len(resumen), 18+len(resumen)+len(chart_data)], [2, 3])
                        
                        chart = PieChart()
                        chart.title = "Distribución de Alertas"
                        chart.height, chart.width = EXCEL_STYLES["CHART_HEIGHT"], EXCEL_STYLES["CHART_WIDTH"]
                        
                        data_ref = Reference(ws2, min_col=3, max_col=3, min_row=19+len(resumen), max_row=18+len(resumen)+len(chart_data))
                        labels_ref = Reference(ws2, min_col=2, min_row=19+len(resumen), max_row=18+len(resumen)+len(chart_data))
                        
                        chart.add_data(data_ref)
                        chart.set_categories(labels_ref)
                        
                        chart.dataLabels = DataLabelList()
                        chart.dataLabels.showPercent = True
                        
                        ws2.add_chart(chart, f"E{22+len(resumen)+len(chart_data)}")
                    except: pass
        
        return True
    except Exception as e:
        print(f"⚠️ Error generando Excel: {e}")
        return False