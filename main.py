import os
import re
import base64
import pandas as pd
import boto3
from datetime import datetime, timedelta
from email.utils import parsedate_to_datetime
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from google.auth.transport.requests import Request
import yagmail

# Importar configuración
from config import AWS_CONFIG, GMAIL_CONFIG, EMAIL_CONFIG, REPORT_CONFIG, EXCEL_STYLES

# Variables globales desde la configuración
HORAS_CUSTOM = REPORT_CONFIG["HORAS_CUSTOM"]

def setup_aws():
    try:
        # Primero configuramos una sesión con credenciales por defecto
        # (del entorno, perfil o rol de instancia EC2)
        boto3.setup_default_session(region_name=AWS_CONFIG["REGION"])
        
        # Creamos un cliente STS para asumir el rol
        sts_client = boto3.client('sts')
        
        # Asumimos el rol especificado en la configuración
        assumed_role = sts_client.assume_role(
            RoleArn=AWS_CONFIG["ROLE_ARN"],
            RoleSessionName=AWS_CONFIG["ROLE_SESSION_NAME"]
        )
        
        # Obtenemos las credenciales temporales
        credentials = assumed_role['Credentials']
        
        # Configuramos la sesión con las credenciales del rol asumido
        boto3.setup_default_session(
            aws_access_key_id=credentials['AccessKeyId'],
            aws_secret_access_key=credentials['SecretAccessKey'],
            aws_session_token=credentials['SessionToken'],
            region_name=AWS_CONFIG["REGION"]
        )
        print(f"✅ Rol AWS asumido: {AWS_CONFIG['ROLE_ARN']}")
        
        # Obtener nombres de cuentas
        try:
            client = boto3.client('organizations', region_name=AWS_CONFIG["REGION"])  
            accounts = []
            for page in client.get_paginator('list_accounts').paginate():
                accounts.extend(page['Accounts'])
            return {a['Id']: a['Name'] for a in accounts if a['Status'] == 'ACTIVE'}
        except Exception as e:
            print(f"⚠️ No se pudieron obtener cuentas de Organizations: {e}")
            # Fallback a cuenta actual
            try:
                sts = boto3.client('sts', region_name=AWS_CONFIG["REGION"])
                account_id = sts.get_caller_identity().get('Account')
                return {account_id: f"Cuenta {account_id}"}
            except Exception as e:
                print(f"⚠️ No se pudo obtener la cuenta actual: {e}")
                return {}
    except Exception as e:
        print(f"❌ Error al asumir el rol AWS: {e}")
        return {}

def setup_gmail():
    creds = None
    # Define paths for credential files
    token_path = GMAIL_CONFIG["TOKEN_PATH"]
    credentials_path = GMAIL_CONFIG["CREDENTIALS_PATH"]
    
    # Ensure credential directory exists
    os.makedirs(os.path.dirname(token_path), exist_ok=True)
    
    if os.path.exists(token_path):
        creds = Credentials.from_authorized_user_file(token_path, GMAIL_CONFIG["SCOPES"])
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(credentials_path, GMAIL_CONFIG["SCOPES"])
            creds = flow.run_local_server(port=0)
        with open(token_path, 'w') as token:
            token.write(creds.to_json())
    return build('gmail', 'v1', credentials=creds)

def get_emails(service, keyword, desde_hora=None):
    query = f"subject:{keyword}"
    if desde_hora:
        query += f" after:{desde_hora.strftime('%Y/%m/%d')}"
    return service.users().messages().list(userId='me', q=query).execute().get('messages', [])

def analizar_mensajes(service, messages, account_names, horas=None):
    data = []
    ahora = datetime.now()

    for msg in messages:
        msg_data = service.users().messages().get(userId='me', id=msg['id'], format='full').execute()
        headers = {h['name']: h['value'] for h in msg_data['payload']['headers']}
        subject = headers.get('Subject', '')
        fecha_raw = headers.get('Date', '')

        try:
            fecha_dt = parsedate_to_datetime(fecha_raw).astimezone()
            if horas and fecha_dt < ahora - timedelta(hours=horas):
                continue
            fecha_str = fecha_dt.strftime('%Y-%m-%d %H:%M:%S')
        except:
            fecha_str = fecha_raw
            fecha_dt = None

        # Extraer cuerpo del mensaje
        payload = msg_data['payload']
        if 'parts' in payload:
            parts = [p for p in payload['parts'] if p['mimeType'] == 'text/plain']
            body = base64.urlsafe_b64decode(parts[0]['body']['data']).decode() if parts else ''
        else:
            body = base64.urlsafe_b64decode(payload['body']['data']).decode()

        # Clasificar alerta
        estado = 'Critica' if 'critical' in subject.lower() else 'Warning' if 'warning' in subject.lower() else 'Informativo' if 'info' in subject.lower() else 'Desconocido'
        
        # Extraer valores
        aws_account = re.search(r'AWS Account:\s+(\d+)', body)
        metric_name = re.search(r'MetricName:\s+([^\s,]+)', body)
        region = re.search(r'Region:\s+([^\s,]+)', body)
        namespace = re.search(r'Namespace:\s+([^\s,]+)', body)
        reason = re.search(r'NewStateReason:\s+([^\n]+)', body)
        
        account_id = aws_account.group(1) if aws_account else ''
        
        data.append({
            'Id cuenta': account_id,
            'Nombre cuenta': account_names.get(account_id, 'Desconocido'),
            'Metrica': metric_name.group(1) if metric_name else '',
            'Namespace': namespace.group(1) if namespace else '',
            'Estado': estado,
            'Fecha': fecha_dt,
            'Fecha_str': fecha_str
        })

    return pd.DataFrame(data)

def aplicar_formato(ws, filas, columnas, titulo=None):
    # Bordes y alineación
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))
    
    for row in ws.iter_rows(min_row=filas[0], max_row=filas[1], min_col=columnas[0], max_col=columnas[1]):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Encabezados - solo la primera fila
    for col in range(columnas[0], columnas[1] + 1):
        cell = ws.cell(row=filas[0], column=col)
        cell.font = Font(bold=True, color=EXCEL_STYLES["HEADER_FONT_COLOR"])
        cell.fill = PatternFill(start_color=EXCEL_STYLES["HEADER_COLOR"], 
                              end_color=EXCEL_STYLES["HEADER_COLOR"], fill_type="solid")
    
    # Ancho columnas
    for col in range(columnas[0], columnas[1] + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    
    # Título
    if titulo:
        ws.merge_cells(start_row=filas[0]-2, start_column=columnas[0], 
                      end_row=filas[0]-2, end_column=columnas[1])
        cell = ws.cell(row=filas[0]-2, column=columnas[0])
        cell.value = titulo
        cell.font = Font(bold=True, size=14)
        cell.alignment = Alignment(horizontal='center')

def generar_recomendaciones(df):
    if len(df) == 0:
        return ["No hay datos suficientes para generar recomendaciones."]
    
    recomendaciones = []
    
    # Alertas críticas
    criticas = df[df['Estado'] == 'Critica']
    if len(criticas) > 0:
        # Métricas problemáticas
        top_metricas = criticas.groupby('Metrica').size().sort_values(ascending=False).head(2)
        for metrica, count in top_metricas.items():
            recomendaciones.append(f"⚠️ La métrica '{metrica}' ha generado {count} alertas críticas.")
        
        # Cuentas con problemas
        top_cuentas = criticas.groupby('Nombre cuenta').size().sort_values(ascending=False).head(1)
        for cuenta, count in top_cuentas.items():
            recomendaciones.append(f"⚠️ La cuenta '{cuenta}' tiene {count} alertas críticas.")
    else:
        recomendaciones.append("✅ No se detectaron alertas críticas en el período.")
    
    return recomendaciones

def generar_excel(df, resumen, periodo, horas=None):
    try:
        sufijo = f"_ultimas_{horas}h" if horas else ""
        archivo = f'{REPORT_CONFIG["EXCEL_DIR"]}/Alertas_{periodo}{sufijo}.xlsx'
        
        # Asegurar que el directorio de Excel existe
        os.makedirs(REPORT_CONFIG["EXCEL_DIR"], exist_ok=True)

        with pd.ExcelWriter(archivo, engine='openpyxl') as writer:
            # Detalle
            if not df.empty:
                columnas = ['Id cuenta', 'Nombre cuenta', 'Metrica', 'Estado', 'Fecha_str']
                for col in ['Región', 'Motivo']:
                    if col in df.columns and df[col].notna().any():
                        columnas.append(col)
                
                df_detalle = df[columnas].rename(columns={
                    'Id cuenta': 'ID Cuenta', 
                    'Nombre cuenta': 'Nombre Cuenta', 
                    'Metrica': 'Métrica',
                    'Fecha_str': 'Fecha'
                })
                
                df_detalle.to_excel(writer, sheet_name='Detalle', index=False)
                aplicar_formato(writer.sheets['Detalle'], [1, len(df_detalle) + 1], [1, len(df_detalle.columns)])
            else:
                pd.DataFrame({'Mensaje': ['No hay alertas en el período seleccionado']}).to_excel(
                    writer, sheet_name='Detalle', index=False)
            
            # Resumen con Dashboard
            pd.DataFrame().to_excel(writer, sheet_name='Resumen', index=False)
            ws2 = writer.sheets['Resumen']
            
            # KPIs
            total = len(df)
            criticas = len(df[df['Estado'] == 'Critica'])
            warnings = len(df[df['Estado'] == 'Warning'])
            info = len(df[df['Estado'] == 'Informativo'])
            
            # Porcentajes
            pct_criticas = (criticas / total * 100) if total > 0 else 0
            pct_warnings = (warnings / total * 100) if total > 0 else 0
            pct_info = (info / total * 100) if total > 0 else 0
            
            # Dashboard
            ws2['B2'] = "DASHBOARD DE ALERTAS"
            ws2['B2'].font = Font(bold=True, size=16)
            ws2.merge_cells('B2:F2')
            ws2['B2'].alignment = Alignment(horizontal='center')
            
            # KPIs
            ws2['B4'], ws2['C4'], ws2['D4'], ws2['E4'] = "Total Alertas", "Críticas", "Warnings", "Informativas"
            ws2['B5'], ws2['C5'], ws2['D5'], ws2['E5'] = total, criticas, warnings, info
            ws2['C6'], ws2['D6'], ws2['E6'] = f"{pct_criticas:.1f}%", f"{pct_warnings:.1f}%", f"{pct_info:.1f}%"
            
            # Formato KPIs
            ws2['B5'].font = Font(bold=True, size=14)
            ws2['C5'].font = Font(bold=True, size=14, color=EXCEL_STYLES["CRITICAL_COLOR"])
            ws2['D5'].font = Font(bold=True, size=14, color=EXCEL_STYLES["WARNING_COLOR"])
            ws2['E5'].font = Font(bold=True, size=14, color=EXCEL_STYLES["INFO_COLOR"])
            
            for col in ['B', 'C', 'D', 'E']:
                ws2[f'{col}4'].font = Font(bold=True)
                for row in range(4, 7):
                    ws2.cell(row=row, column=ord(col)-64).alignment = Alignment(horizontal='center')
                    ws2.cell(row=row, column=ord(col)-64).border = Border(
                        left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin')
                    )
            
            # Recomendaciones
            ws2.cell(row=8, column=2, value="RECOMENDACIONES:")
            ws2.cell(row=8, column=2).font = Font(bold=True)
            
            for i, rec in enumerate(generar_recomendaciones(df)[:3]):
                ws2.cell(row=9+i, column=2, value=rec)
            
            row_start = 13  # Después de recomendaciones
            
            # Análisis por servicio
            if not df.empty and 'Namespace' in df.columns and df['Namespace'].notna().any():
                try:
                    servicios = df.groupby(['Namespace', 'Estado']).size().unstack(fill_value=0).reset_index()
                    servicios = servicios.rename(columns={'Namespace': 'Servicio AWS'})
                    
                    for estado in ['Critica', 'Warning', 'Informativo']:
                        if estado not in servicios.columns:
                            servicios[estado] = 0
                    
                    servicios['Total'] = servicios[['Critica', 'Warning', 'Informativo']].sum(axis=1)
                    servicios = servicios.sort_values('Total', ascending=False)
                    
                    # Título
                    ws2.cell(row=row_start, column=2, value="ANÁLISIS POR SERVICIO AWS")
                    ws2.cell(row=row_start, column=2).font = Font(bold=True, size=14)
                    ws2.merge_cells(start_row=row_start, start_column=2, end_row=row_start, end_column=6)
                    ws2.cell(row=row_start, column=2).alignment = Alignment(horizontal='center')
                    
                    # Encabezados
                    headers = ['Servicio AWS', 'Critica', 'Warning', 'Informativo', 'Total']
                    for j, header in enumerate(headers):
                        ws2.cell(row=row_start+2, column=2+j, value=header)
                    
                    # Datos
                    for i, row in enumerate(servicios.itertuples(index=False)):
                        for j, val in enumerate(row):
                            ws2.cell(row=row_start+3+i, column=2+j, value=val)
                    
                    aplicar_formato(ws2, [row_start+2, row_start+3+len(servicios)-1], 
                                   [2, 2+len(headers)-1])
                    
                    row_start = row_start + 5 + len(servicios)
                except Exception as e:
                    print(f"Error en análisis de servicios: {e}")
            
            # Resumen por cuenta
            if not resumen.empty:
                # Título
                ws2.cell(row=row_start, column=2, value="RESUMEN POR CUENTA Y MÉTRICA")
                ws2.cell(row=row_start, column=2).font = Font(bold=True, size=14)
                ws2.merge_cells(start_row=row_start, start_column=2, end_row=row_start, end_column=6)
                ws2.cell(row=row_start, column=2).alignment = Alignment(horizontal='center')
                
                # Encabezados
                headers = ['ID Cuenta', 'Nombre Cuenta', 'Métrica', 'Critica', 'Warning', 'Informativo', 'Total']
                for j, header in enumerate(headers):
                    ws2.cell(row=row_start+2, column=2+j, value=header)
                
                # Datos
                for i, row in enumerate(resumen.itertuples(index=False)):
                    for j, val in enumerate(row):
                        ws2.cell(row=row_start+3+i, column=2+j, value=val)
                
                aplicar_formato(ws2, [row_start+2, row_start+3+len(resumen)-1], 
                               [2, 2+len(headers)-1])
                
                # Gráfico circular
                if all(col in resumen.columns for col in ['Critica', 'Warning', 'Informativo']):
                    try:
                        # Datos para gráfico - solo incluir valores > 0
                        chart_data = []
                        if criticas > 0:
                            chart_data.append(("Críticas", criticas))
                        if warnings > 0:
                            chart_data.append(("Warnings", warnings))
                        if info > 0:
                            chart_data.append(("Informativas", info))
                        
                        # Escribir encabezados
                        ws2.cell(row=row_start+len(resumen)+5, column=2, value="Estado")
                        ws2.cell(row=row_start+len(resumen)+5, column=3, value="Cantidad")
                        
                        # Escribir datos
                        for i, (label, value) in enumerate(chart_data):
                            ws2.cell(row=row_start+len(resumen)+6+i, column=2, value=label)
                            ws2.cell(row=row_start+len(resumen)+6+i, column=3, value=value)
                        
                        # Aplicar formato a la tabla
                        aplicar_formato(ws2, 
                                      [row_start+len(resumen)+5, row_start+len(resumen)+5+len(chart_data)], 
                                      [2, 3])
                        
                        # Gráfico
                        if criticas + warnings + info > 0:
                            try:
                                crear_grafico_circular(ws2, row_start, resumen, chart_data)
                            except Exception as e:
                                print(f"Error en gráfico: {e}")
                    except Exception as e:
                        print(f"Error en datos para gráfico: {e}")
        
        return True
    except Exception as e:
        print(f"Error al generar Excel: {e}")
        return False

def crear_grafico_circular(ws, row_start, resumen, chart_data):
    # Crear un gráfico circular simple
    chart = PieChart()
    chart.title = "Distribución de Alertas por Severidad"
    chart.style = EXCEL_STYLES["CHART_STYLE"]
    chart.height = EXCEL_STYLES["CHART_HEIGHT"]  # Altura en unidades
    chart.width = EXCEL_STYLES["CHART_WIDTH"]   # Ancho en unidades
    
    # Referencias a los datos y etiquetas
    data_ref = Reference(ws, min_col=3, max_col=3, 
                       min_row=row_start+len(resumen)+6, 
                       max_row=row_start+len(resumen)+5+len(chart_data))
    labels_ref = Reference(ws, min_col=2, 
                         min_row=row_start+len(resumen)+6, 
                         max_row=row_start+len(resumen)+5+len(chart_data))
    
    # Configurar el gráfico
    chart.add_data(data_ref)
    chart.set_categories(labels_ref)
    
    # Mostrar porcentajes en las etiquetas
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showPercent = True
    chart.dataLabels.showVal = False
    chart.dataLabels.showCatName = False
    
    # Calcular posición relativa para la gráfica
    chart_position = row_start + len(resumen) + len(chart_data) + 10
    
    # Añadir el gráfico a la hoja en posición relativa
    ws.add_chart(chart, f"E{chart_position}")

def crear_mensaje_correo(periodo, horas, df):
    # Personalizar asunto según el periodo
    fecha_actual = datetime.now().strftime('%Y-%m-%d')
    if periodo == 'mensual':
        subject = f"Concentrado mensual de alarmas Estafeta: {fecha_actual}"
        detalle = f"Se adjunta el concentrado mensual de alarmas de Estafeta.\n\n"
    elif periodo == 'custom_horas':
        subject = f"Concentrado de alarmas Estafeta: {len(df)} alertas en últimas {horas}h ({fecha_actual})"
        detalle = f"Se adjunta el concentrado de alarmas de Estafeta de las últimas {horas} horas.\n"
    else:
        subject = f"Concentrado de alarmas Estafeta: {len(df)} alertas - {periodo.capitalize()} ({fecha_actual})"
        detalle = f"Se adjunta el concentrado {periodo} de alarmas de Estafeta.\n"
    
    # Construir mensaje
    message = "Buen día,\n\n"
    message += detalle
    
    if periodo != 'mensual':
        message += f"Se detectaron {len(df)} alertas en este periodo.\n\n"
        
    message += "\nEste es un mensaje automático. \n"
    message += "Por favor, no responda a este correo.\n\n"
    message += "Saludos cordiales."
    
    return subject, message

def generar_reporte(service, keyword, periodo='diario', horas=None):
    print(f"\n📊 Generando reporte: {periodo.upper()} ({horas or 'Días'})")

    try:
        account_names = setup_aws()
        
        # Rango de fechas
        desde = datetime.now()
        if horas:
            desde -= timedelta(hours=horas)
        elif periodo == 'semanal':
            desde -= timedelta(days=7)
        elif periodo == 'mensual':
            desde -= timedelta(days=30)
        else:
            desde -= timedelta(days=1)

        # Obtener y analizar mensajes
        messages = get_emails(service, keyword, desde)
        df = analizar_mensajes(service, messages, account_names, horas)

        # Crear resumen
        resumen = pd.DataFrame()
        if not df.empty:
            try:
                resumen = (
                    df.groupby(['Id cuenta', 'Nombre cuenta', 'Metrica', 'Estado'])
                    .size().reset_index(name='Cantidad')
                    .pivot_table(index=['Id cuenta', 'Nombre cuenta', 'Metrica'], 
                                columns='Estado', values='Cantidad', fill_value=0)
                    .reset_index()
                )
                
                # Asegurar columnas de estado
                for estado in ['Critica', 'Warning', 'Informativo']:
                    if estado not in resumen.columns:
                        resumen[estado] = 0
                        
                # Total
                resumen['Total'] = resumen[['Critica', 'Warning', 'Informativo']].sum(axis=1)
                resumen = resumen.sort_values('Total', ascending=False)
            except Exception as e:
                print(f"Error al crear resumen: {e}")

        # Generar Excel
        if generar_excel(df, resumen, periodo, horas):
            print(f"✅ {len(df)} alertas exportadas")
        else:
            print("⚠️ Problemas al generar el reporte")
            return
    except Exception as e:
        print(f"❌ Error: {e}")
        return

    # Enviar correo con el reporte
    try:
        sufijo = f"_ultimas_{horas}h" if horas else ""
        nombre_archivo = f'Alertas_{periodo}{sufijo}.xlsx'
        archivo_adjunto = os.path.join(os.getcwd(), REPORT_CONFIG["EXCEL_DIR"], nombre_archivo)
        
        # Crear asunto y mensaje
        subject, message = crear_mensaje_correo(periodo, horas, df)
        
        # Enviar correo
        yag = yagmail.SMTP(EMAIL_CONFIG["REMITENTE"], EMAIL_CONFIG["PASSWORD"])
        yag.send(
            to=EMAIL_CONFIG["DESTINATARIO"],
            subject=subject,
            contents=message,
            cc=EMAIL_CONFIG["COPIAS"],
            attachments=archivo_adjunto
        )
        print(f"✅ Reporte enviado a {EMAIL_CONFIG['DESTINATARIO']} y {len(EMAIL_CONFIG['COPIAS'])} destinatarios en copia")
        print(f"✅ Archivo adjunto: {archivo_adjunto}")
    except Exception as e:
        print(f"⚠️ Error al enviar el correo: {e}")

def main(periodos, keyword=REPORT_CONFIG["DEFAULT_KEYWORD"]):
    print("\n📈 === MONITOREO Y PERSPECTIVA AVANZADA ===")
    
    # Autenticar
    service = setup_gmail()
    
    # Generar reportes para cada periodo especificado
    for periodo in periodos:
        horas = HORAS_CUSTOM if periodo == 'custom_horas' else None
        generar_reporte(service, keyword, periodo, horas)

    print("\n📁 Reportes generados con éxito. Incluyen:")
    print("  ✅ Resumen con KPIs principales integrados")
    print("  ✅ Análisis por servicio AWS en la misma hoja")
    print("  ✅ Gráficos visuales de distribución de alertas")
    print("  ✅ Recomendaciones automáticas")

def actualizar_horas_custom(horas):
    global HORAS_CUSTOM
    HORAS_CUSTOM = horas

if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description='Genera reportes de alertas de AWS')
    parser.add_argument('--periodos', nargs='+', 
                      default=["mensual"],
                      choices=["diario", "semanal", "mensual", "custom_horas", "all"],
                      help='Periodos a generar (diario, semanal, mensual, custom_horas)')
    parser.add_argument('--keyword', default=REPORT_CONFIG["DEFAULT_KEYWORD"],
                      help='Palabra clave para buscar en los correos')
    parser.add_argument('--horas', type=int, default=REPORT_CONFIG["HORAS_CUSTOM"],
                      help='Número de horas para el periodo custom_horas')
    
    args = parser.parse_args()
    
    # Procesar periodos
    periodos = ["diario", "semanal", "mensual", "custom_horas"] if "all" in args.periodos else args.periodos
    
    # Actualizar horas para periodo personalizado
    if "custom_horas" in periodos:
        actualizar_horas_custom(args.horas)
    
    main(periodos, args.keyword)